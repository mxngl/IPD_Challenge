from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .models import ImpactVector


DEFAULT_TEMPLATE_PATH = (
    Path(__file__).resolve().parents[2] / "STV_Template" / "STV_ConceptA_Bambo.xlsx"
)

ASSEMBLY_COLUMN_MAP = {
    "Foundation": 2,
    "Interior Wall": 3,
    "Exterior Wall": 4,
    "Floor": 5,
    "Roof": 6,
    "Window": 7,
    "Columns": 8,
    "Beams": 9,
    "MEP": 10,
    "Energy": 11,
    "Misc": 12,
}

ASSEMBLY_ALIASES = {
    "Columns": "Column",
    "Beams": "Beam",
}


@dataclass(slots=True)
class MaterialRecord:
    assembly: str
    material_type: str
    embodied_total: ImpactVector
    materials: ImpactVector
    transport: ImpactVector
    construction: ImpactVector
    unit_multiplier: float


@dataclass(slots=True)
class TeamFactors:
    team: str
    grid_electricity: ImpactVector
    target_carbon_factor: float
    target_water_factor: float
    target_energy_factor: float


@dataclass(slots=True)
class FuelRecord:
    fuel_type: str
    water: float
    ozone: float
    mj_per_fu: float
    carbon_per_mj: float


class STVReferenceData:
    def __init__(
        self,
        materials: dict[tuple[str, str], MaterialRecord],
        materials_by_name: dict[str, list[MaterialRecord]],
        valid_materials: dict[str, set[str]],
        teams: dict[str, TeamFactors],
        fuels: dict[str, FuelRecord],
    ) -> None:
        self.materials = materials
        self.materials_by_name = materials_by_name
        self.valid_materials = valid_materials
        self.teams = teams
        self.fuels = fuels

    @classmethod
    def from_workbook(cls, workbook_path: Path | str | None = None) -> "STVReferenceData":
        path = Path(workbook_path) if workbook_path else DEFAULT_TEMPLATE_PATH
        wb = load_workbook(path, data_only=True)

        materials: dict[tuple[str, str], MaterialRecord] = {}
        materials_by_name: dict[str, list[MaterialRecord]] = {}

        ws_lca = wb["LCA Data"]
        for row in range(8, 108):
            assembly = ws_lca[f"B{row}"].value
            material_type = ws_lca[f"C{row}"].value
            if not assembly or not material_type:
                continue
            record = MaterialRecord(
                assembly=str(assembly),
                material_type=str(material_type),
                embodied_total=ImpactVector(
                    carbon=float(ws_lca[f"D{row}"].value or 0.0),
                    energy=float(ws_lca[f"E{row}"].value or 0.0),
                    water=float(ws_lca[f"F{row}"].value or 0.0),
                    ozone=float(ws_lca[f"G{row}"].value or 0.0),
                ),
                materials=ImpactVector(
                    carbon=float(ws_lca[f"H{row}"].value or 0.0),
                    energy=float(ws_lca[f"I{row}"].value or 0.0),
                    water=float(ws_lca[f"J{row}"].value or 0.0),
                    ozone=float(ws_lca[f"K{row}"].value or 0.0),
                ),
                transport=ImpactVector(
                    carbon=float(ws_lca[f"L{row}"].value or 0.0),
                    energy=float(ws_lca[f"M{row}"].value or 0.0),
                    water=float(ws_lca[f"N{row}"].value or 0.0),
                    ozone=float(ws_lca[f"O{row}"].value or 0.0),
                ),
                construction=ImpactVector(
                    carbon=float(ws_lca[f"P{row}"].value or 0.0),
                    energy=float(ws_lca[f"Q{row}"].value or 0.0),
                    water=float(ws_lca[f"R{row}"].value or 0.0),
                    ozone=float(ws_lca[f"S{row}"].value or 0.0),
                ),
                unit_multiplier=float(ws_lca[f"T{row}"].value or 1.0),
            )
            key = (record.assembly, record.material_type)
            materials[key] = record
            materials_by_name.setdefault(record.material_type, []).append(record)

        ws_lists = wb["Lists"]
        valid_materials: dict[str, set[str]] = {}
        for assembly, column_index in ASSEMBLY_COLUMN_MAP.items():
            allowed: set[str] = set()
            for row in range(2, 252):
                value = ws_lists.cell(row=row, column=column_index).value
                if value is None:
                    continue
                text = str(value).strip()
                if text:
                    allowed.add(text)
            valid_materials[assembly] = allowed

        for assembly, material_type in materials:
            list_assembly = {
                "Column": "Columns",
                "Beam": "Beams",
            }.get(assembly, assembly)
            valid_materials.setdefault(list_assembly, set()).add(material_type)

        ws_team = wb["LCA Data"]
        teams: dict[str, TeamFactors] = {}
        for row in range(115, 122):
            team = ws_team[f"B{row}"].value
            if not team:
                continue
            teams[str(team)] = TeamFactors(
                team=str(team),
                grid_electricity=ImpactVector(
                    carbon=float(ws_team[f"C{row}"].value or 0.0),
                    energy=float(ws_team[f"H{row}"].value or 0.0),
                    water=float(ws_team[f"I{row}"].value or 0.0),
                    ozone=float(ws_team[f"J{row}"].value or 0.0),
                ),
                target_carbon_factor=float(ws_team[f"K{row}"].value or 0.0),
                target_water_factor=float(ws_team[f"L{row}"].value or 0.0),
                target_energy_factor=float(ws_team[f"M{row}"].value or 0.0),
            )

        ws_fuels = wb["Cogen Data"]
        fuels: dict[str, FuelRecord] = {}
        for row in range(5, 15):
            fuel_type = ws_fuels[f"B{row}"].value
            if not fuel_type:
                continue
            fuels[str(fuel_type)] = FuelRecord(
                fuel_type=str(fuel_type),
                water=float(ws_fuels[f"F{row}"].value or 0.0),
                ozone=float(ws_fuels[f"G{row}"].value or 0.0),
                mj_per_fu=float(ws_fuels[f"I{row}"].value or 0.0),
                carbon_per_mj=float(ws_fuels[f"J{row}"].value or 0.0),
            )

        return cls(
            materials=materials,
            materials_by_name=materials_by_name,
            valid_materials=valid_materials,
            teams=teams,
            fuels=fuels,
        )

    def validate_item(self, assembly: str, material_type: str) -> None:
        if assembly not in self.valid_materials:
            known = ", ".join(sorted(self.valid_materials))
            raise ValueError(f"Unknown assembly '{assembly}'. Expected one of: {known}")
        if material_type not in self.valid_materials[assembly]:
            allowed = ", ".join(sorted(self.valid_materials[assembly]))
            raise ValueError(
                f"Material '{material_type}' is not valid for assembly '{assembly}'. "
                f"Allowed values: {allowed}"
            )

    def get_material(self, assembly: str, material_type: str) -> MaterialRecord:
        self.validate_item(assembly, material_type)
        canonical_assembly = ASSEMBLY_ALIASES.get(assembly, assembly)
        try:
            return self.materials[(canonical_assembly, material_type)]
        except KeyError as exc:
            raise ValueError(
                f"No LCA data found for assembly '{assembly}' and material '{material_type}'."
            ) from exc

    def get_team(self, team: str) -> TeamFactors:
        try:
            return self.teams[team]
        except KeyError as exc:
            known = ", ".join(sorted(self.teams))
            raise ValueError(f"Unknown team '{team}'. Expected one of: {known}") from exc

    def get_fuel(self, fuel_type: str) -> FuelRecord:
        try:
            return self.fuels[fuel_type]
        except KeyError as exc:
            known = ", ".join(sorted(self.fuels))
            raise ValueError(f"Unknown fuel '{fuel_type}'. Expected one of: {known}") from exc
